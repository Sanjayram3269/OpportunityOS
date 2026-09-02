"""Excel workbook builder — generates openpyxl workbook from export data.

Produces a multi-sheet workbook with:
- Opportunities, Companies, Leads, Outreach, FollowUps
- Interactions, Evidence, Campaigns, Summary
- Headers, autofilter, frozen panes, sensible widths
- ISO-8601 timezone-preserving date formatting
- Clickable hyperlinks for URLs
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


# ── Formatting constants ─────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_MAX_TEXT_WIDTH = 50
_MIN_COL_WIDTH = 10
_DATE_FORMAT = "%Y-%m-%d %H:%M:%S %Z"


def _safe_value(val: Any) -> Any:
    """Convert value for Excel cell, handling None and datetime.

    Datetimes are formatted as ISO-8601 strings with timezone preserved.
    """
    if val is None:
        return ""
    if isinstance(val, datetime):
        if val.tzinfo is not None:
            return val.strftime("%Y-%m-%dT%H:%M:%S%z")
        return val.strftime("%Y-%m-%dT%H:%M:%S")
    return val


def _write_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    """Write a sheet with headers, data, autofilter, and frozen panes."""
    ws = wb.create_sheet(title=title)

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=_safe_value(val))

    # Autofilter
    if rows:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    # Freeze header row
    ws.freeze_panes = "A2"

    # Column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in rows:
            if col_idx - 1 < len(row):
                val = row[col_idx - 1]
                if val is not None:
                    max_len = max(max_len, min(len(str(val)), _MAX_TEXT_WIDTH))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            _MIN_COL_WIDTH, min(max_len + 2, _MAX_TEXT_WIDTH + 2)
        )


def _write_sheet_with_urls(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    url_columns: list[str] | None = None,
) -> None:
    """Write a sheet with optional hyperlink support for URL columns."""
    _write_sheet(wb, title, headers, rows)

    if not url_columns or not rows:
        return

    ws = wb[title]
    url_col_indices = [
        headers.index(col) for col in url_columns if col in headers
    ]

    for row_idx in range(2, len(rows) + 2):
        for col_idx in url_col_indices:
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=cell.value)
                cell.style = "Hyperlink"


def build_workbook(data: dict[str, Any]) -> BytesIO:
    """Build an Excel workbook from pre-structured export data.

    Args:
        data: Dictionary with keys for each sheet:
            - opportunities: (headers, rows)
            - companies: (headers, rows)
            - leads: (headers, rows)
            - outreach: (headers, rows)
            - followups: (headers, rows)
            - interactions: (headers, rows)
            - evidence: (headers, rows)
            - campaigns: (headers, rows)
            - summary: (headers, rows)

    Returns:
        BytesIO containing the .xlsx file.
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Opportunities — with URL hyperlink on Source URL
    _write_sheet_with_urls(
        wb, "Opportunities",
        *data.get("opportunities", ([], [])),
        url_columns=["Source URL"],
    )

    # Companies — with URL hyperlinks
    _write_sheet_with_urls(
        wb, "Companies",
        *data.get("companies", ([], [])),
        url_columns=["Website", "LinkedIn URL"],
    )

    # Leads — with URL hyperlinks
    _write_sheet_with_urls(
        wb, "Leads",
        *data.get("leads", ([], [])),
        url_columns=["LinkedIn URL", "Website URL"],
    )

    _write_sheet(wb, "Outreach", *data.get("outreach", ([], [])))
    _write_sheet(wb, "FollowUps", *data.get("followups", ([], [])))
    _write_sheet(wb, "Interactions", *data.get("interactions", ([], [])))
    _write_sheet(wb, "Evidence", *data.get("evidence", ([], [])))
    _write_sheet(wb, "Campaigns", *data.get("campaigns", ([], [])))
    _write_sheet(wb, "App Timeline", *data.get("application_timeline", ([], [])))
    _write_sheet(wb, "Summary", *data.get("summary", ([], [])))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
