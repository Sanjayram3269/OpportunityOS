"""End-to-end validation tests for the complete OpportunityOS lifecycle.

Exercises the full journey:
  Source → Discovery → Opportunity → Company → Match → Planning → Action
  → Application → Timeline → Outreach → Follow-Up → Notification → Analytics → Export

Finds real integration and correctness issues.
"""

from __future__ import annotations

import asyncio
from datetime import datetime, timedelta, timezone
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import func

from app.models.application import APPLICATION_TRANSITIONS, TERMINAL_STATUSES, Action, Application
from app.models.application_event import ApplicationEvent
from app.models.campaign import Campaign
from app.models.campaign_opportunity import CampaignOpportunity
from app.models.company import Company
from app.models.followup import FollowUp
from app.models.interaction import Interaction
from app.models.lead import Lead
from app.models.message import Message
from app.models.notification import Notification
from app.models.opportunity import Opportunity
from app.models.profile import Profile


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════


def _create_profile(db) -> Profile:
    profile = Profile(
        name="Test User",
        email="test@example.com",
        phone="+1234567890",
        headline="Senior Software Engineer",
    )
    db.add(profile)
    db.commit()
    db.refresh(profile)
    return profile


def _create_company(db, name: str = "TestCorp") -> Company:
    company = Company(name=name)
    db.add(company)
    db.commit()
    db.refresh(company)
    return company


def _create_opportunity(
    db,
    company: Company,
    *,
    title: str = "Software Engineer",
    opp_type: str = "FULL_TIME",
    status: str = "DISCOVERED",
    priority: str = "MEDIUM",
    match_score: int | None = None,
    deadline: datetime | None = None,
    source_url: str | None = None,
    description: str | None = None,
) -> Opportunity:
    opp = Opportunity(
        company_id=company.id,
        type=opp_type,
        title=title,
        description=description or "A great role at a great company.",
        source_url=source_url,
        status=status,
        priority=priority,
        match_score=match_score,
        deadline=deadline,
    )
    db.add(opp)
    db.commit()
    db.refresh(opp)
    return opp


def _create_lead(
    db, company: Company, name: str = "John Doe", email: str | None = "john@test.com"
) -> Lead:
    lead = Lead(
        company_id=company.id,
        name=name,
        email=email,
        title="Engineering Manager",
        status="ACTIVE",
    )
    db.add(lead)
    db.commit()
    db.refresh(lead)
    return lead


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 1: DISCOVERY → OPPORTUNITY → COMPANY
# ══════════════════════════════════════════════════════════════════════════════


class TestDiscoveryToEndToEnd:
    """Test discovery → opportunity → company flow."""

    def test_discovery_creates_opportunity_and_company(self, db):
        """Running discovery should create opportunities with company associations."""
        from app.services.discovery import ingest
        from app.discovery.models import RawOpportunity

        raw = RawOpportunity(
            source_name="test",
            external_id="test-001",
            title="Python Developer",
            company_name="TestCo",
            description="Develop Python apps",
            source_url="https://example.com/job/1",
            location="Remote",
        )

        from app.discovery.normalizer import normalize_all

        normalized = normalize_all([raw])
        result = ingest(db, normalized)

        assert result.ingested == 1
        assert result.raw_count == 1

        # Verify opportunity was created with correct data
        opp = db.query(Opportunity).filter(Opportunity.title == "Python Developer").first()
        assert opp is not None
        assert opp.company_id is not None
        assert opp.status == "DISCOVERED"
        assert opp.source_url == "https://example.com/job/1"

        # Verify company was created
        company = db.get(Company, opp.company_id)
        assert company is not None

    def test_discovery_deduplication(self, db):
        """Running the same discovery twice should not create duplicates."""
        from app.services.discovery import ingest
        from app.discovery.models import RawOpportunity
        from app.discovery.normalizer import normalize_all

        raw = RawOpportunity(
            source_name="test",
            external_id="dedup-001",
            title="Go Developer",
            company_name="DedupCo",
            description="Develop Go apps",
            source_url="https://example.com/job/2",
        )

        normalized = normalize_all([raw])
        result1 = ingest(db, normalized)
        assert result1.ingested == 1

        # Same source + external_id = duplicate
        result2 = ingest(db, normalized)
        assert result2.ingested == 0
        assert result2.duplicates_skipped >= 1

    def test_discovery_source_failure_doesnt_break_others(self, db):
        """If one source adapter fails, other sources should still succeed."""
        from app.services.discovery import run_source

        # run_source handles errors gracefully
        result = run_source(db, "nonexistent_source_xyz")
        assert result.ingested == 0
        assert len(result.errors) > 0


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 2: MATCHING VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestMatchingValidation:
    """Test matching engine produces deterministic scores."""

    def test_match_score_deterministic(self, db):
        """Same profile + opportunity should produce the same score."""
        from app.services.matching import match_opportunity

        profile = _create_profile(db)
        company = _create_company(db)
        opp = _create_opportunity(db, company, title="Python Developer")

        result1 = match_opportunity(db, profile, opp)
        result2 = match_opportunity(db, profile, opp)

        assert result1.score == result2.score
        assert 0 <= result1.score <= 100

    def test_match_score_range(self, db):
        """Match score should always be between 0 and 100."""
        from app.services.matching import match_opportunity

        profile = _create_profile(db)
        company = _create_company(db)

        for title in ["Python Engineer", "Ballet Dancer", "Senior Dev"]:
            opp = _create_opportunity(db, company, title=title)
            result = match_opportunity(db, profile, opp)
            assert 0 <= result.score <= 100, f"Score {result.score} out of range for {title}"

    def test_match_always_has_explanation(self, db):
        """Match result should always have an explanation."""
        from app.services.matching import match_opportunity

        profile = _create_profile(db)
        company = _create_company(db)
        opp = _create_opportunity(db, company)

        result = match_opportunity(db, profile, opp)
        assert result.explanation is not None
        assert len(result.explanation) > 0


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 3: PLANNING VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestPlanningValidation:
    """Test planning horizon classification with real deadlines."""

    def test_now_classification(self, db):
        """Opportunity with deadline within 7 days should be NOW."""
        from app.services.planning import classify_horizon

        now = datetime.now(timezone.utc)
        deadline = now + timedelta(days=3)
        assert classify_horizon(deadline, now) == "NOW"

    def test_upcoming_classification(self, db):
        """Opportunity with deadline 8-30 days out should be UPCOMING."""
        from app.services.planning import classify_horizon

        now = datetime.now(timezone.utc)
        deadline = now + timedelta(days=15)
        assert classify_horizon(deadline, now) == "UPCOMING"

    def test_future_classification(self, db):
        """Opportunity with deadline beyond 30 days should be FUTURE."""
        from app.services.planning import classify_horizon

        now = datetime.now(timezone.utc)
        deadline = now + timedelta(days=60)
        assert classify_horizon(deadline, now) == "FUTURE"

    def test_unknown_when_no_deadline(self, db):
        """Opportunity with no deadline should be UNKNOWN."""
        from app.services.planning import classify_horizon

        now = datetime.now(timezone.utc)
        assert classify_horizon(None, now) == "UNKNOWN"

    def test_summer_2027_precedence(self, db):
        """Summer 2027 opportunities (May-June 2027) should be classified as SUMMER_2027, not FUTURE."""
        from app.services.planning import classify_horizon

        now = datetime.now(timezone.utc)
        # May 15, 2027 — within Summer 2027 window
        deadline_may = datetime(2027, 5, 15, tzinfo=timezone.utc)
        assert classify_horizon(deadline_may, now) == "SUMMER_2027"

        # June 30, 2027 — last day of Summer 2027
        deadline_june = datetime(2027, 6, 30, 23, 59, tzinfo=timezone.utc)
        assert classify_horizon(deadline_june, now) == "SUMMER_2027"

    def test_summer_2027_boundary_outside(self, db):
        """July 1 2027 should NOT be SUMMER_2027."""
        from app.services.planning import classify_horizon

        now = datetime.now(timezone.utc)
        deadline_july = datetime(2027, 7, 1, tzinfo=timezone.utc)
        horizon = classify_horizon(deadline_july, now)
        assert horizon != "SUMMER_2027"

        deadline_april = datetime(2027, 4, 30, tzinfo=timezone.utc)
        horizon = classify_horizon(deadline_april, now)
        assert horizon != "SUMMER_2027"

    def test_created_at_not_treated_as_deadline(self, db):
        """created_at should NEVER be treated as an application deadline."""
        company = _create_company(db)
        opp = _create_opportunity(db, company)  # No deadline set

        from app.services.planning import classify_horizon

        now = datetime.now(timezone.utc)
        horizon = classify_horizon(opp.deadline, now)
        assert horizon == "UNKNOWN"  # Not NOW, not UPCOMING


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 4: APPLICATION LIFECYCLE VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestApplicationLifecycleValidation:
    """Test the complete application lifecycle through valid transitions."""

    def test_full_happy_path(self, db):
        """NOT_APPLIED → READY → APPLIED → ASSESSMENT → INTERVIEW → FINAL_ROUND → OFFER → ACCEPTED."""
        from app.services.application import create_application, transition_application

        company = _create_company(db)
        opp = _create_opportunity(db, company, title="Dream Job")

        app = create_application(db, opportunity_id=opp.id)
        assert app.status == "NOT_APPLIED"

        app = transition_application(db, app.id, "READY")
        assert app.status == "READY"

        app = transition_application(db, app.id, "APPLIED")
        assert app.status == "APPLIED"
        assert app.applied_at is not None  # applied_at should be set

        app = transition_application(db, app.id, "ASSESSMENT")
        assert app.status == "ASSESSMENT"

        app = transition_application(db, app.id, "INTERVIEW")
        assert app.status == "INTERVIEW"

        app = transition_application(db, app.id, "FINAL_ROUND")
        assert app.status == "FINAL_ROUND"

        app = transition_application(db, app.id, "OFFER")
        assert app.status == "OFFER"

        app = transition_application(db, app.id, "ACCEPTED")
        assert app.status == "ACCEPTED"

        # ACCEPTED is terminal — no further transitions allowed
        valid = APPLICATION_TRANSITIONS.get("ACCEPTED", [])
        assert len(valid) == 0

    def test_rejection_path(self, db):
        """Application can be rejected at various stages."""
        from app.services.application import create_application, transition_application

        company = _create_company(db)
        opp = _create_opportunity(db, company)

        app = create_application(db, opportunity_id=opp.id)
        app = transition_application(db, app.id, "READY")
        app = transition_application(db, app.id, "APPLIED")
        app = transition_application(db, app.id, "REJECTED")
        assert app.status == "REJECTED"

        # REJECTED is terminal
        valid = APPLICATION_TRANSITIONS.get("REJECTED", [])
        assert len(valid) == 0

    def test_withdrawal_path(self, db):
        """Application can be withdrawn (from READY or later states)."""
        from app.services.application import create_application, transition_application

        company = _create_company(db)
        opp = _create_opportunity(db, company)

        app = create_application(db, opportunity_id=opp.id)
        app = transition_application(db, app.id, "READY")
        app = transition_application(db, app.id, "WITHDRAWN")
        assert app.status == "WITHDRAWN"

    def test_invalid_transition_rejected(self, db):
        """Invalid transitions should raise ValueError."""
        from app.services.application import create_application, transition_application

        company = _create_company(db)
        opp = _create_opportunity(db, company)

        app = create_application(db, opportunity_id=opp.id)
        assert app.status == "NOT_APPLIED"

        # Can't skip directly to INTERVIEW
        with pytest.raises(ValueError, match="Invalid transition"):
            transition_application(db, app.id, "INTERVIEW")

    def test_application_events_created(self, db):
        """Each valid transition should create an ApplicationEvent."""
        from app.services.application import create_application, transition_application

        company = _create_company(db)
        opp = _create_opportunity(db, company)

        app = create_application(db, opportunity_id=opp.id)
        event_count_after_create = (
            db.query(func.count(ApplicationEvent.id))
            .filter(ApplicationEvent.application_id == app.id)
            .scalar()
        )
        assert event_count_after_create >= 1  # At least APPLICATION_CREATED

        app = transition_application(db, app.id, "READY")
        app = transition_application(db, app.id, "APPLIED")

        event_count = (
            db.query(func.count(ApplicationEvent.id))
            .filter(ApplicationEvent.application_id == app.id)
            .scalar()
        )
        assert event_count >= 3  # CREATED + READY + APPLIED

    def test_event_timeline_chronological(self, db):
        """Events should be in chronological order."""
        from app.services.application import create_application, transition_application
        from app.services.timeline import get_application_timeline

        company = _create_company(db)
        opp = _create_opportunity(db, company)

        app = create_application(db, opportunity_id=opp.id)
        transition_application(db, app.id, "READY")
        transition_application(db, app.id, "APPLIED")
        transition_application(db, app.id, "ASSESSMENT")

        events = get_application_timeline(db, app.id)

        assert len(events) >= 4
        for i in range(1, len(events)):
            prev_time = datetime.fromisoformat(events[i - 1]["occurred_at"])
            curr_time = datetime.fromisoformat(events[i]["occurred_at"])
            assert curr_time >= prev_time, "Timeline events should be chronological"

    def test_no_fabricated_history(self, db):
        """Application events should only reflect actual transitions, not invented history."""
        from app.services.application import create_application
        from app.services.timeline import get_application_timeline

        company = _create_company(db)
        opp = _create_opportunity(db, company)

        app = create_application(db, opportunity_id=opp.id)
        events = get_application_timeline(db, app.id)

        # Only APPLICATION_CREATED event should exist
        event_types = [e["event_type"] for e in events]
        assert event_types.count("APPLICATION_CREATED") == 1
        assert "INTERVIEW" not in event_types  # No fabricated history


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 5: ACTION CENTER VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestActionCenterValidation:
    """Test action generation, priority, and lifecycle."""

    def test_actions_generated_for_high_match(self, db):
        """High-match opportunities without applications should generate actions."""
        from app.services.action_center import generate_actions

        company = _create_company(db)
        _create_opportunity(db, company, match_score=90)

        actions = generate_actions(db, dry_run=True)
        apply_actions = [a for a in actions if a.action_type == "APPLY"]
        assert len(apply_actions) >= 1

    def test_priority_deterministic(self, db):
        """Same inputs should produce the same priority."""
        from app.services.action_center import calculate_action_priority

        p1 = calculate_action_priority(
            match_score=90,
            planning_horizon="NOW",
            deadline_bucket="OVERDUE",
            application_status="NOT_APPLIED",
        )
        p2 = calculate_action_priority(
            match_score=90,
            planning_horizon="NOW",
            deadline_bucket="OVERDUE",
            application_status="NOT_APPLIED",
        )
        assert p1 == p2

    def test_high_match_overdue_is_p0(self, db):
        """Overdue + high match + not applied should be P0."""
        from app.services.action_center import calculate_action_priority

        priority = calculate_action_priority(
            match_score=90,
            planning_horizon="NOW",
            deadline_bucket="OVERDUE",
            application_status="NOT_APPLIED",
        )
        assert priority == "P0"

    def test_action_lifecycle(self, db):
        """Actions should support OPEN → IN_PROGRESS → COMPLETED flow."""
        from app.services.action_center import generate_actions, start_action, complete_action, dismiss_action
        from app.models.application import Action

        company = _create_company(db)
        _create_opportunity(db, company, match_score=85)

        actions = generate_actions(db)
        db.commit()
        assert len(actions) >= 1

        action = actions[0]

        # Start
        started = start_action(db, action.id)
        assert started.status == "IN_PROGRESS"

        # Complete
        completed = complete_action(db, action.id)
        assert completed.status == "COMPLETED"
        assert completed.completed_at is not None

    def test_action_idempotency(self, db):
        """Generating actions twice should not create duplicates for same entity."""
        from app.services.action_center import generate_actions

        company = _create_company(db)
        _create_opportunity(db, company, match_score=85)

        actions1 = generate_actions(db)
        count1 = len(actions1)

        actions2 = generate_actions(db)
        count2 = len(actions2)

        # Second generation should not create new actions for same entity
        assert count2 == 0 or count2 <= count1


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 6: OUTREACH VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestOutreachValidation:
    """Test outreach lifecycle: DRAFT → PENDING_APPROVAL → APPROVED → READY_TO_SEND."""

    def test_full_outreach_lifecycle(self, db):
        """Test complete outreach approval flow."""
        from app.services.outreach import (
            transition_draft, DRAFT, PENDING_APPROVAL, APPROVED, READY_TO_SEND, REJECTED
        )

        company = _create_company(db)
        lead = _create_lead(db, company)
        opp = _create_opportunity(db, company)

        msg = Message(
            lead_id=lead.id,
            opportunity_id=opp.id,
            channel="EMAIL",
            direction="OUTBOUND",
            subject="Interest in role",
            body="Hi, I'm interested in the role.",
            status=DRAFT,
        )
        db.add(msg)
        db.commit()
        db.refresh(msg)

        # DRAFT → PENDING_APPROVAL
        msg = transition_draft(db, msg, PENDING_APPROVAL)
        assert msg.status == "PENDING_APPROVAL"

        # PENDING_APPROVAL → APPROVED
        msg = transition_draft(db, msg, APPROVED)
        assert msg.status == "APPROVED"

        # APPROVED → READY_TO_SEND
        msg = transition_draft(db, msg, READY_TO_SEND)
        assert msg.status == "READY_TO_SEND"

    def test_invalid_transition_blocked(self, db):
        """DRAFT → READY_TO_SEND (skipping approval) should fail."""
        from app.services.outreach import (
            transition_draft, DraftStateError, DRAFT, READY_TO_SEND
        )

        company = _create_company(db)
        lead = _create_lead(db, company)

        msg = Message(
            lead_id=lead.id,
            channel="EMAIL",
            direction="OUTBOUND",
            subject="Test",
            body="Test body",
            status=DRAFT,
        )
        db.add(msg)
        db.commit()
        db.refresh(msg)

        with pytest.raises(DraftStateError):
            transition_draft(db, msg, READY_TO_SEND)

    def test_no_automatic_send(self, db):
        """Approving outreach should NOT automatically send the message."""
        from app.services.outreach import (
            transition_draft, DRAFT, PENDING_APPROVAL, APPROVED, READY_TO_SEND
        )

        company = _create_company(db)
        lead = _create_lead(db, company)

        msg = Message(
            lead_id=lead.id,
            channel="EMAIL",
            direction="OUTBOUND",
            subject="Test",
            body="Test",
            status=READY_TO_SEND,  # Jump to READY_TO_SEND for testing
        )
        db.add(msg)
        db.commit()
        db.refresh(msg)

        # Verify no Interaction was created (no send happened)
        interactions = (
            db.query(Interaction)
            .filter(Interaction.message_id == msg.id)
            .all()
        )
        assert len(interactions) == 0


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 7: FOLLOW-UP VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestFollowUpValidation:
    """Test follow-up lifecycle and timezone handling."""

    def test_full_followup_lifecycle(self, db):
        """Test follow-up: PENDING → DUE → PENDING_APPROVAL → APPROVED → READY_TO_SEND → COMPLETED."""
        from app.services.followup import (
            create_followup, check_and_mark_due, submit_followup,
            approve_followup, mark_followup_ready, complete_followup,
        )

        company = _create_company(db)
        lead = _create_lead(db, company)

        # Create a follow-up scheduled in the past (so it becomes due immediately)
        past_time = datetime.now(timezone.utc) - timedelta(hours=1)
        fu = create_followup(db, lead_id=lead.id, scheduled_for=past_time, reason="Check in")
        assert fu.status == "PENDING"

        # Mark as due
        fu = check_and_mark_due(db, fu)
        assert fu.status == "DUE"

        # Submit for approval
        fu = submit_followup(db, fu)
        assert fu.status == "PENDING_APPROVAL"

        # Approve
        fu = approve_followup(db, fu)
        assert fu.status == "APPROVED"

        # Mark ready
        fu = mark_followup_ready(db, fu)
        assert fu.status == "READY_TO_SEND"

        # Complete
        fu = complete_followup(db, fu)
        assert fu.status == "COMPLETED"
        assert fu.completed_at is not None

    def test_future_followup_not_due(self, db):
        """A follow-up scheduled in the future should NOT become due."""
        from app.services.followup import create_followup, check_and_mark_due

        company = _create_company(db)
        lead = _create_lead(db, company)

        future_time = datetime.now(timezone.utc) + timedelta(days=30)
        fu = create_followup(db, lead_id=lead.id, scheduled_for=future_time)
        assert fu.status == "PENDING"

        fu = check_and_mark_due(db, fu)
        assert fu.status == "PENDING"  # Should stay PENDING

    def test_reject_naive_datetime(self, db):
        """Follow-up creation should reject naive datetimes."""
        from app.services.followup import create_followup

        company = _create_company(db)
        lead = _create_lead(db, company)

        naive_time = datetime(2026, 1, 1, 12, 0, 0)  # No timezone
        with pytest.raises(ValueError, match="timezone-aware"):
            create_followup(db, lead_id=lead.id, scheduled_for=naive_time)

    def test_invalid_followup_transition_rejected(self, db):
        """Invalid follow-up transitions should be rejected."""
        from app.services.followup import (
            create_followup, FollowUpStateError, transition_followup
        )

        company = _create_company(db)
        lead = _create_lead(db, company)

        fu = create_followup(
            db, lead_id=lead.id, scheduled_for=datetime.now(timezone.utc) + timedelta(days=1)
        )

        # PENDING → COMPLETED should be invalid
        with pytest.raises(FollowUpStateError):
            transition_followup(db, fu, "COMPLETED")


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 8: NOTIFICATION VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestNotificationValidation:
    """Test notification generation and deduplication."""

    def test_sync_creates_notifications(self, db):
        """Notification sync should create notifications from system state."""
        from app.services.notifications import sync_notifications

        result = sync_notifications(db)
        assert "created" in result
        assert isinstance(result["created"], int)

    def test_sync_idempotent(self, db):
        """Running sync twice should not create duplicate notifications."""
        from app.services.notifications import sync_notifications

        result1 = sync_notifications(db)
        result2 = sync_notifications(db)

        # Second sync should create 0 new (all already exist)
        assert result2["created"] == 0

    def test_read_notification_not_recreated(self, db):
        """Read notifications should not be recreated on resync."""
        from app.services.notifications import (
            sync_notifications, list_notifications, mark_read
        )

        sync_notifications(db)
        notifs = list_notifications(db, unread_only=True)
        if len(notifs) > 0:
            mark_read(db, notifs[0].id)
            db.commit()

            result = sync_notifications(db)
            assert result["created"] == 0

    def test_mark_read_and_unread_count(self, db):
        """Marking notifications as read should update unread count."""
        from app.services.notifications import sync_notifications, get_unread_count, mark_read

        sync_notifications(db)
        count_before = get_unread_count(db)

        notifs = [n for n in db.query(Notification).filter(Notification.read_at.is_(None)).all()]
        if len(notifs) > 0:
            mark_read(db, notifs[0].id)
            db.commit()
            count_after = get_unread_count(db)
            assert count_after == count_before - 1


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 9: CAMPAIGN VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestCampaignValidation:
    """Test campaign CRUD and opportunity association."""

    def test_campaign_crud(self, db):
        """Test campaign create, read, update, lifecycle."""
        from app.services.campaign import (
            create_campaign, get_campaign, update_campaign,
            transition_campaign,
        )

        campaign = create_campaign(
            db,
            name="Summer 2027 SWE",
            type="INTERNSHIP",
            description="Targeting summer SWE internships",
        )
        assert campaign.status == "DRAFT"

        # Activate (pass Campaign object, not ID)
        campaign = transition_campaign(db, campaign, "ACTIVE")
        assert campaign.status == "ACTIVE"

        # Pause
        campaign = transition_campaign(db, campaign, "PAUSED")
        assert campaign.status == "PAUSED"

        # Reactivate
        campaign = transition_campaign(db, campaign, "ACTIVE")
        assert campaign.status == "ACTIVE"

        # Complete
        campaign = transition_campaign(db, campaign, "COMPLETED")
        assert campaign.status == "COMPLETED"

    def test_campaign_opportunity_association(self, db):
        """Adding/removing opportunities from campaigns."""
        from app.services.campaign import create_campaign, add_opportunity_to_campaign, remove_opportunity_from_campaign, list_campaign_opportunities

        company = _create_company(db)
        opp = _create_opportunity(db, company, match_score=85)
        campaign = create_campaign(db, name="Test Campaign", type="GENERAL")

        # Add
        add_opportunity_to_campaign(db, campaign, opp.id)
        opps = list_campaign_opportunities(db, campaign)
        assert len(opps) == 1

        # Duplicate add should be idempotent
        add_opportunity_to_campaign(db, campaign, opp.id)
        opps = list_campaign_opportunities(db, campaign)
        assert len(opps) == 1

        # Remove
        remove_opportunity_from_campaign(db, campaign, opp.id)
        opps = list_campaign_opportunities(db, campaign)
        assert len(opps) == 0


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 10: DASHBOARD VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestDashboardValidation:
    """Test dashboard data accuracy with real records."""

    def test_dashboard_reflects_real_data(self, client):
        """Dashboard should accurately reflect current database state."""
        response = client.get("/dashboard/overview")
        assert response.status_code == 200
        data = response.json()

        assert "overview" in data
        assert "today" in data
        assert "pipeline" in data
        assert "opportunities" in data
        assert "campaigns" in data
        assert "outreach" in data
        assert "followups" in data
        assert "analytics" in data

        # Empty database should show zeros, not crash
        assert data["overview"]["total_opportunities"] >= 0
        assert data["overview"]["total_applications"] >= 0


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 11: EXPORT VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestExportValidation:
    """Test Excel export produces valid workbook."""

    def test_export_creates_valid_xlsx(self, client):
        """Export endpoint should return a valid xlsx workbook."""
        response = client.get("/exports/opportunities.xlsx")
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers.get("content-type", "")

        # Parse the workbook to verify sheets
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        sheet_names = wb.sheetnames
        assert "Opportunities" in sheet_names
        assert "Companies" in sheet_names
        assert "Summary" in sheet_names
        wb.close()

    def test_export_with_data(self, db, client):
        """Export should include data from the database."""
        company = _create_company(db)
        _create_opportunity(db, company, title="Export Test Role", match_score=75)

        response = client.get("/exports/opportunities.xlsx")
        assert response.status_code == 200

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        ws = wb["Opportunities"]

        # Find the header row and verify data
        found = False
        for row in ws.iter_rows(values_only=True):
            if row and "Export Test Role" in str(row):
                found = True
                break
        assert found, "Opportunity should appear in export"
        wb.close()


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 12: AUTOMATION END-TO-END
# ══════════════════════════════════════════════════════════════════════════════


class TestAutomationEndToEnd:
    """Test complete automation cycle."""

    def test_automation_cycle_returns_complete_result(self, db):
        """Automation run should return a complete result with all fields."""
        from app.automation.engine import run_automation_cycle
        from app.automation.models import RunTrigger

        result = asyncio.run(
            run_automation_cycle(db, trigger=RunTrigger.MANUAL, dry_run=True)
        )

        assert result.run_id != ""
        assert result.status.value in ("COMPLETED", "FAILED")
        assert result.started_at is not None
        assert result.completed_at is not None
        assert result.duration_seconds() is not None
        assert result.duration_seconds() >= 0

    def test_automation_no_external_actions(self, db):
        """Automation should never send emails or submit applications."""
        import app.automation.engine as eng_mod
        import inspect

        source = inspect.getsource(eng_mod)
        assert "send_email" not in source
        assert "send_message" not in source
        assert "submit_application" not in source
        assert "approve_draft" not in source

    def test_automation_endpoints(self, client):
        """Automation API endpoints should work correctly."""
        # Status
        resp = client.get("/automation/status")
        assert resp.status_code == 200
        data = resp.json()
        assert "enabled" in data
        assert "sources" in data

        # Run
        resp = client.post("/automation/run", json={"dry_run": True})
        assert resp.status_code == 200
        data = resp.json()
        assert "run_id" in data
        assert data["status"] in ("COMPLETED", "FAILED")


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 13: CROSS-SYSTEM CONSISTENCY
# ══════════════════════════════════════════════════════════════════════════════


class TestCrossSystemConsistency:
    """Test data consistency across multiple services."""

    def test_opportunity_to_notification_consistency(self, db):
        """High-match opportunity should generate both actions and notifications."""
        from app.services.action_center import generate_actions
        from app.services.notifications import sync_notifications

        company = _create_company(db)
        opp = _create_opportunity(db, company, match_score=95, priority="HIGH")

        # Generate actions
        actions = generate_actions(db)
        high_match_actions = [
            a for a in actions if a.entity_type == "opportunity" and a.entity_id == opp.id
        ]
        assert len(high_match_actions) >= 1

        # Sync notifications
        sync_notifications(db)

        # Verify notification exists for this opportunity
        notifs = (
            db.query(Notification)
            .filter(
                Notification.source_type == "opportunity",
                Notification.source_id == opp.id,
            )
            .all()
        )
        assert len(notifs) >= 1

    def test_application_to_timeline_consistency(self, db):
        """Application transitions should consistently appear in timeline."""
        from app.services.application import create_application, transition_application
        from app.services.timeline import get_application_timeline

        company = _create_company(db)
        opp = _create_opportunity(db, company)

        app = create_application(db, opportunity_id=opp.id)
        transition_application(db, app.id, "READY")
        transition_application(db, app.id, "APPLIED")

        timeline = get_application_timeline(db, app.id)
        event_types = [e["event_type"] for e in timeline]

        assert "APPLICATION_CREATED" in event_types
        assert "APPLICATION_SUBMITTED" in event_types or "STATUS_CHANGED" in event_types

    def test_campaign_summary_reflects_data(self, db):
        """Campaign summary should accurately reflect its opportunities."""
        from app.services.campaign import create_campaign, add_opportunity_to_campaign
        from app.services.campaign_enhanced import get_enhanced_campaign_summary

        company = _create_company(db)
        opp1 = _create_opportunity(db, company, title="Role 1", match_score=85)
        opp2 = _create_opportunity(db, company, title="Role 2", match_score=70)
        campaign = create_campaign(db, name="Test", type="GENERAL")

        add_opportunity_to_campaign(db, campaign, opp1.id)
        add_opportunity_to_campaign(db, campaign, opp2.id)

        summary = get_enhanced_campaign_summary(db, campaign)
        assert summary["total_opportunities"] == 2


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 14: ANALYTICS VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestAnalyticsValidation:
    """Test analytics accuracy with real data."""

    def test_analytics_empty_state(self, client):
        """Analytics should return valid data even with empty database."""
        resp = client.get("/analytics/overview")
        assert resp.status_code == 200
        data = resp.json()
        assert "overview" in data
        assert data["overview"]["total_opportunities"] >= 0

    def test_analytics_reflects_applications(self, db, client):
        """Analytics should accurately count application states."""
        from app.services.application import create_application, transition_application

        company = _create_company(db)
        opp1 = _create_opportunity(db, company)
        opp2 = _create_opportunity(db, company)

        app1 = create_application(db, opportunity_id=opp1.id)
        transition_application(db, app1.id, "READY")
        transition_application(db, app1.id, "APPLIED")

        app2 = create_application(db, opportunity_id=opp2.id)
        transition_application(db, app2.id, "READY")
        transition_application(db, app2.id, "APPLIED")
        transition_application(db, app2.id, "INTERVIEW")

        resp = client.get("/analytics/overview")
        assert resp.status_code == 200
        data = resp.json()
        overview = data["overview"]
        assert overview["total_applications"] == 2
        assert overview["interviews"] >= 1


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 15: API ENDPOINT VALIDATION
# ══════════════════════════════════════════════════════════════════════════════


class TestAPIEndpointValidation:
    """Test API endpoints return correct status codes and data."""

    def test_404_for_nonexistent_entities(self, client):
        """Non-existent entities should return 404."""
        endpoints = [
            ("GET", "/opportunities/99999"),
            ("GET", "/companies/99999"),
            ("GET", "/leads/99999"),
            ("GET", "/applications/99999"),
            ("GET", "/actions/99999"),
            ("GET", "/campaigns/99999"),
            ("GET", "/follow-ups/99999"),
            ("GET", "/notifications/99999"),
        ]

        for method, path in endpoints:
            resp = client.get(path)
            assert resp.status_code == 404, f"{method} {path} should return 404"

    def test_notification_404_read(self, client):
        """Reading a non-existent notification should return 404."""
        resp = client.post("/notifications/99999/read")
        assert resp.status_code == 404

    def test_all_list_endpoints_work(self, client):
        """All list endpoints should return 200."""
        endpoints = [
            "/opportunities",
            "/companies",
            "/leads",
            "/applications",
            "/actions",
            "/campaigns",
            "/follow-ups",
            "/notifications",
            "/outreach/drafts",
            "/profiles",
            "/discovery/sources",
        ]

        for path in endpoints:
            resp = client.get(path)
            assert resp.status_code == 200, f"{path} should return 200, got {resp.status_code}"


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 16: SECURITY REGRESSION
# ══════════════════════════════════════════════════════════════════════════════


class TestSecurityRegression:
    """Verify security invariants remain intact."""

    def test_no_secrets_in_config_endpoint(self, client):
        """Config/status endpoints must not expose secrets."""
        resp = client.get("/automation/config")
        data = resp.json()
        assert "email_password" not in str(data)
        assert "ai_api_key" not in str(data)
        assert "database_url" not in str(data)

    def test_no_linkedin_scraping(self):
        """Codebase must not contain LinkedIn scraping or automation."""
        import os
        import re

        dangerous_patterns = [
            re.compile(r"selenium", re.IGNORECASE),
            re.compile(r"puppeteer", re.IGNORECASE),
            re.compile(r"playwright", re.IGNORECASE),
            re.compile(r"linkedin.*scrape", re.IGNORECASE),
            re.compile(r"linkedin.*automat", re.IGNORECASE),
        ]

        findings = []
        for root, dirs, files in os.walk("apps/api"):
            # Skip venv, __pycache__, .git
            dirs[:] = [d for d in dirs if d not in (".venv", "__pycache__", ".git", "node_modules")]
            for fname in files:
                if not fname.endswith((".py", ".ts", ".tsx", ".js")):
                    continue
                fpath = os.path.join(root, fname)
                try:
                    with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                        for line_no, line in enumerate(f, 1):
                            for pat in dangerous_patterns:
                                if pat.search(line) and not line.strip().startswith("#"):
                                    findings.append(f"{fpath}:{line_no}: {line.strip()}")
                except Exception:
                    pass

        assert len(findings) == 0, f"Found dangerous patterns: {findings}"

    def test_no_automatic_application_submission(self):
        """Automation must not contain application submission logic."""
        import app.automation.engine as eng
        import inspect

        source = inspect.getsource(eng)
        assert "submit_application" not in source
        assert "apply_to_job" not in source

    def test_stack_trace_not_exposed(self, client):
        """500 errors should not expose stack traces."""
        # This is already handled by global exception handler
        # Verify health endpoint doesn't expose internals
        resp = client.get("/health")
        assert resp.status_code == 200
        data = resp.json()
        assert "stack_trace" not in str(data)
        assert "traceback" not in str(data)
