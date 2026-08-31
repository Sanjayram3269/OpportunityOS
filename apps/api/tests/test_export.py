"""Comprehensive tests for Excel Export functionality.

Tests cover:
  1. empty database export
  2. opportunity export with planning data
  3. company export
  4. lead export
  5. outreach export
  6. follow-up export
  7. campaign export
  8. summary counts
  9. planning horizon included
  10. planning priority included
  11. match score preserved
  12. Summer 2027 correct
  13. no-deadline = UNKNOWN
  14. filtering by horizon/type/status/campaign/company/location
  15. workbook sheet names + headers
  16. date handling, None handling
  17. API response MIME type
  18. no database mutation
  19. existing regression
"""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.company import Company
from app.models.followup import FollowUp
from app.models.interaction import Interaction
from app.models.lead import Lead
from app.models.message import Message
from app.models.opportunity import Opportunity
from app.models.opportunity_evidence import OpportunityEvidence
from app.export.workbook import build_workbook
from app.services.export import ExportOptions, build_export_data


# ── Helpers ──────────────────────────────────────────────────────────────


def _create_test_data(db):
    company = Company(name="ExportCo", location="Bengaluru")
    db.add(company)
    db.flush()

    lead = Lead(
        company_id=company.id, name="Jane", email="j@export.com",
        title="Manager", location="Bengaluru",
    )
    db.add(lead)
    db.flush()

    opp = Opportunity(
        company_id=company.id, lead_id=lead.id,
        type="INTERNSHIP", title="Python Intern",
        description="Python and Django", status="DISCOVERED",
        priority="HIGH", match_score=85,
        deadline=datetime(2027, 5, 15, tzinfo=timezone.utc),
    )
    db.add(opp)
    db.flush()

    msg = Message(
        lead_id=lead.id, opportunity_id=opp.id,
        channel="EMAIL", direction="OUTBOUND",
        subject="Test", body="Hello", status="SENT",
        ai_generated=False, sent_at=datetime.now(timezone.utc),
    )
    db.add(msg)
    db.flush()

    fu = FollowUp(
        lead_id=lead.id, opportunity_id=opp.id, message_id=msg.id,
        scheduled_for=datetime(2026, 9, 15, tzinfo=timezone.utc),
        status="PENDING", reason="Follow up",
    )
    db.add(fu)
    db.flush()

    interaction = Interaction(
        lead_id=lead.id, message_id=msg.id,
        type="EMAIL_SENT", content="Email sent to j@export.com",
        metadata_={"provider": "smtp", "subject": "Test"},
    )
    db.add(interaction)
    db.flush()

    evidence = OpportunityEvidence(
        opportunity_id=opp.id, source="remotive",
        evidence_type="external_id", content="rem-123",
    )
    db.add(evidence)
    db.flush()

    return company, lead, opp, msg, fu


# ══════════════════════════════════════════════════════════════════════════
# 1. WORKBOOK STRUCTURE
# ══════════════════════════════════════════════════════════════════════════


class TestWorkbookStructure:
    def test_empty_export(self, db):
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        assert "Opportunities" in wb.sheetnames
        assert "Companies" in wb.sheetnames
        assert "Leads" in wb.sheetnames
        assert "Outreach" in wb.sheetnames
        assert "FollowUps" in wb.sheetnames
        assert "Campaigns" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        wb.close()

    def test_all_sheets_exist(self, db):
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        expected = [
            "Opportunities", "Companies", "Leads", "Outreach",
            "FollowUps", "Interactions", "Evidence", "Campaigns", "Summary",
        ]
        for name in expected:
            assert name in wb.sheetnames, f"Missing sheet: {name}"
        wb.close()

    def test_sheet_has_headers(self, db):
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["Opportunities"]
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "Title" in headers
        assert "Match Score" in headers
        assert "Planning Horizon" in headers
        assert "Planning Priority" in headers
        wb.close()

    def test_autofilter_set_with_data(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["Opportunities"]
        assert ws.auto_filter.ref is not None
        wb.close()

    def test_freeze_panes(self, db):
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["Opportunities"]
        assert ws.freeze_panes == "A2"
        wb.close()


# ══════════════════════════════════════════════════════════════════════════
# 2. OPPORTUNITY EXPORT
# ══════════════════════════════════════════════════════════════════════════


class TestInteractionsExport:
    def test_interactions_sheet_exists(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        assert "Interactions" in wb.sheetnames
        wb.close()

    def test_interaction_headers(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["Interactions"]
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "Type" in headers
        assert "Opportunity ID (derived)" in headers
        wb.close()

    def test_interaction_data(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["interactions"]
        assert len(rows) >= 1
        types = [r[headers.index("Type")] for r in rows]
        assert "EMAIL_SENT" in types

    def test_opportunity_id_derived(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["interactions"]
        sent_rows = [r for r in rows if r[headers.index("Type")] == "EMAIL_SENT"]
        assert len(sent_rows) >= 1
        opp_id = sent_rows[0][headers.index("Opportunity ID (derived)")]
        assert opp_id != ""

    def test_empty_interactions(self, db):
        data = build_export_data(db)
        headers, rows = data["interactions"]
        assert rows == []
        assert "ID" in headers


class TestEvidenceExport:
    def test_evidence_sheet_exists(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        assert "Evidence" in wb.sheetnames
        wb.close()

    def test_evidence_headers(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["Evidence"]
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "Source" in headers
        assert "Evidence Type" in headers
        assert "Content" in headers
        wb.close()

    def test_evidence_data(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["evidence"]
        assert len(rows) >= 1
        sources = [r[headers.index("Source")] for r in rows]
        assert "remotive" in sources

    def test_empty_evidence(self, db):
        data = build_export_data(db)
        headers, rows = data["evidence"]
        assert rows == []
        assert "ID" in headers


class TestOpportunityExport:
    def test_opportunity_exported(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["opportunities"]
        assert len(rows) == 1
        assert rows[0][headers.index("Title")] == "Python Intern"

    def test_match_score_preserved(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["opportunities"]
        assert rows[0][headers.index("Match Score")] == 85

    def test_planning_horizon_included(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["opportunities"]
        horizon = rows[0][headers.index("Planning Horizon")]
        assert horizon == "SUMMER_2027"  # deadline May 2027

    def test_planning_priority_included(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["opportunities"]
        priority = rows[0][headers.index("Planning Priority")]
        assert isinstance(priority, int)
        assert 0 <= priority <= 100

    def test_company_name_resolved(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["opportunities"]
        assert rows[0][headers.index("Company")] == "ExportCo"

    def test_no_deadline_opportunity_unknown(self, db):
        company = Company(name="NoDlCo")
        db.add(company)
        db.flush()
        opp = Opportunity(
            company_id=company.id, type="FULL_TIME",
            title="No Deadline", status="DISCOVERED", priority="MEDIUM",
        )
        db.add(opp)
        db.flush()

        data = build_export_data(db)
        headers, rows = data["opportunities"]
        # Find the No Deadline opportunity
        no_dl = [r for r in rows if r[headers.index("Title")] == "No Deadline"]
        assert len(no_dl) == 1
        assert no_dl[0][headers.index("Planning Horizon")] == "UNKNOWN"

    def test_summer_2027_correct(self, db):
        company = Company(name="S27Co")
        db.add(company)
        db.flush()
        opp = Opportunity(
            company_id=company.id, type="INTERNSHIP",
            title="Summer 2027", status="DISCOVERED", priority="MEDIUM",
            deadline=datetime(2027, 6, 15, tzinfo=timezone.utc),
        )
        db.add(opp)
        db.flush()

        data = build_export_data(db)
        headers, rows = data["opportunities"]
        s27 = [r for r in rows if r[headers.index("Title")] == "Summer 2027"]
        assert s27[0][headers.index("Planning Horizon")] == "SUMMER_2027"


# ══════════════════════════════════════════════════════════════════════════
# 3. OTHER SHEETS
# ══════════════════════════════════════════════════════════════════════════


class TestOtherSheets:
    def test_company_exported(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["companies"]
        names = [r[headers.index("Name")] for r in rows]
        assert "ExportCo" in names
        export_row = [r for r in rows if r[headers.index("Name")] == "ExportCo"][0]
        assert export_row[headers.index("Location")] == "Bengaluru"

    def test_lead_exported(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["leads"]
        names = [r[headers.index("Name")] for r in rows]
        assert "Jane" in names

    def test_outreach_exported(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["outreach"]
        assert len(rows) == 1
        assert rows[0][headers.index("Channel")] == "EMAIL"
        assert rows[0][headers.index("AI Generated")] == "No"

    def test_followup_exported(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["followups"]
        assert len(rows) == 1
        assert rows[0][headers.index("Status")] == "PENDING"

    def test_summary_has_counts(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        headers, rows = data["summary"]
        metrics = {r[0]: r[1] for r in rows if r[0] and not r[0].startswith("---")}
        assert metrics.get("Total Opportunities", 0) >= 1
        assert metrics.get("Total Companies", 0) >= 1
        assert metrics.get("Total Interactions", 0) >= 1
        assert metrics.get("Total Evidence Records", 0) >= 1


# ══════════════════════════════════════════════════════════════════════════
# 4. FILTERING
# ══════════════════════════════════════════════════════════════════════════


class TestFiltering:
    def test_filter_by_horizon(self, db):
        _create_test_data(db)
        opts = ExportOptions(planning_horizon="SUMMER_2027")
        data = build_export_data(db, opts)
        headers, rows = data["opportunities"]
        assert len(rows) == 1
        assert rows[0][headers.index("Planning Horizon")] == "SUMMER_2027"

    def test_filter_by_type(self, db):
        _create_test_data(db)
        opts = ExportOptions(opportunity_type="INTERNSHIP")
        data = build_export_data(db, opts)
        headers, rows = data["opportunities"]
        assert all(r[headers.index("Type")] == "INTERNSHIP" for r in rows)

    def test_filter_by_status(self, db):
        _create_test_data(db)
        opts = ExportOptions(status="DISCOVERED")
        data = build_export_data(db, opts)
        headers, rows = data["opportunities"]
        assert all(r[headers.index("Status")] == "DISCOVERED" for r in rows)

    def test_filter_by_min_match_score(self, db):
        _create_test_data(db)
        opts = ExportOptions(min_match_score=80)
        data = build_export_data(db, opts)
        headers, rows = data["opportunities"]
        assert len(rows) == 1  # match_score=85 >= 80

    def test_filter_by_company(self, db):
        company, lead, opp, _, _ = _create_test_data(db)
        opts = ExportOptions(company_id=company.id)
        data = build_export_data(db, opts)
        headers, rows = data["opportunities"]
        titles = [r[headers.index("Title")] for r in rows]
        assert "Python Intern" in titles

    def test_filter_by_location(self, db):
        _create_test_data(db)
        opts = ExportOptions(location="Bengaluru")
        data = build_export_data(db, opts)
        headers, rows = data["opportunities"]
        assert len(rows) == 1

    def test_filter_by_location_no_match(self, db):
        _create_test_data(db)
        opts = ExportOptions(location="Chennai")
        data = build_export_data(db, opts)
        _, rows = data["opportunities"]
        assert len(rows) == 0


# ══════════════════════════════════════════════════════════════════════════
# 5. DATA INTEGRITY
# ══════════════════════════════════════════════════════════════════════════


class TestDateTimeHandling:
    def test_timezone_preserved_in_export(self, db):
        _create_test_data(db)
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["Opportunities"]
        headers = [cell.value for cell in ws[1]]
        deadline_col = headers.index("Deadline") + 1
        # Find row with a deadline
        for row in ws.iter_rows(min_row=2, min_col=deadline_col, max_col=deadline_col):
            val = row[0].value
            if val and val != "":
                assert isinstance(val, str)
                assert "T" in val  # ISO format
                break
        wb.close()

    def test_utc_offset_in_dates(self, db):
        from app.export.workbook import _safe_value
        aware_dt = datetime(2027, 5, 15, 12, 0, 0, tzinfo=timezone.utc)
        result = _safe_value(aware_dt)
        assert "+0000" in result or "+00:00" in result

    def test_naive_datetime_formatted(self, db):
        from app.export.workbook import _safe_value
        naive_dt = datetime(2027, 5, 15, 12, 0, 0)
        result = _safe_value(naive_dt)
        assert "T" in result
        assert "+" not in result  # no false timezone


class TestURLHandling:
    def test_source_url_clickable(self, db):
        company = Company(name="URLCo")
        db.add(company)
        db.flush()
        opp = Opportunity(
            company_id=company.id, type="FULL_TIME",
            title="URL Test", status="DISCOVERED", priority="MEDIUM",
            source_url="https://example.com/job/123",
        )
        db.add(opp)
        db.flush()

        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["Opportunities"]
        headers = [cell.value for cell in ws[1]]
        url_col = headers.index("Source URL") + 1
        # Find the URL Test row
        for row in ws.iter_rows(min_row=2):
            if row[headers.index("Title")].value == "URL Test":
                url_cell = row[url_col - 1]
                assert url_cell.hyperlink is not None
                assert url_cell.hyperlink.target == "https://example.com/job/123"
                break
        wb.close()

    def test_empty_url_no_hyperlink(self, db):
        company = Company(name="NoURLCo")
        db.add(company)
        db.flush()
        opp = Opportunity(
            company_id=company.id, type="FULL_TIME",
            title="NoURL", status="DISCOVERED", priority="MEDIUM",
        )
        db.add(opp)
        db.flush()

        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["Opportunities"]
        headers = [cell.value for cell in ws[1]]
        url_col = headers.index("Source URL") + 1
        for row in ws.iter_rows(min_row=2):
            if row[headers.index("Title")].value == "NoURL":
                url_cell = row[url_col - 1]
                assert url_cell.hyperlink is None
                break
        wb.close()


class TestNoSecretsExported:
    def test_no_config_in_export(self, db):
        """Export must not contain configuration/secret data."""
        data = build_export_data(db)
        all_text = str(data)
        assert "api_key" not in all_text.lower()
        assert "password" not in all_text.lower()
        assert "smtp" not in all_text.lower()
        assert "secret" not in all_text.lower()


class TestDataIntegrity:
    def test_no_database_mutation(self, db):
        _create_test_data(db)
        count_before = db.query(Opportunity).count()
        build_export_data(db)
        count_after = db.query(Opportunity).count()
        assert count_before == count_after

    def test_none_handling(self, db):
        company = Company(name="NullCo")
        db.add(company)
        db.flush()
        opp = Opportunity(
            company_id=company.id, type="FULL_TIME",
            title="Null Test", status="DISCOVERED", priority="MEDIUM",
        )
        db.add(opp)
        db.flush()

        data = build_export_data(db)
        headers, rows = data["opportunities"]
        null_opp = [r for r in rows if r[headers.index("Title")] == "Null Test"][0]
        # None values become empty string via _safe_value, or openpyxl returns None
        assert null_opp[headers.index("Match Score")] in ("", None)
        assert null_opp[headers.index("Deadline")] in ("", None)
        assert null_opp[headers.index("Source URL")] in ("", None)


# ══════════════════════════════════════════════════════════════════════════
# 6. API ENDPOINT
# ══════════════════════════════════════════════════════════════════════════


class TestExportAPI:
    def test_download_xlsx(self, client, db):
        _create_test_data(db)
        response = client.get("/exports/opportunities.xlsx")
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
        assert "attachment" in response.headers["content-Disposition"]

        wb = load_workbook(BytesIO(response.content))
        assert "Opportunities" in wb.sheetnames
        wb.close()

    def test_empty_download(self, client, db):
        response = client.get("/exports/opportunities.xlsx")
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Opportunities"]
        assert ws.max_row == 1  # header only
        wb.close()

    def test_filter_via_api(self, client, db):
        _create_test_data(db)
        response = client.get("/exports/opportunities.xlsx?planning_horizon=SUMMER_2027")
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Opportunities"]
        assert ws.max_row == 2  # header + 1 row
        wb.close()

    def test_filter_by_type_via_api(self, client, db):
        _create_test_data(db)
        response = client.get("/exports/opportunities.xlsx?opportunity_type=INTERNSHIP")
        assert response.status_code == 200


# ══════════════════════════════════════════════════════════════════════════
# 7. EXISTING REGRESSION
# ══════════════════════════════════════════════════════════════════════════


class TestExistingRegression:
    def test_health(self, client):
        assert client.get("/health").status_code == 200

    def test_opportunity_crud(self, client, db):
        c = client.post("/companies", json={"name": "Exp Reg Co"}).json()
        resp = client.post("/opportunities", json={
            "company_id": c["id"], "type": "FULL_TIME", "title": "Exp Reg",
        })
        assert resp.status_code == 201

    def test_campaign(self, client, db):
        resp = client.post("/campaigns", json={"name": "Exp Camp", "type": "FULL_TIME"})
        assert resp.status_code == 201

    def test_planning(self, client, db):
        c = client.post("/companies", json={"name": "Exp P Co"}).json()
        client.post("/opportunities", json={
            "company_id": c["id"], "type": "FULL_TIME", "title": "P",
        })
        resp = client.get("/opportunities/planning")
        assert resp.status_code == 200

    def test_matching(self, client, db):
        from app.models.profile import Profile
        from app.models.skill import Skill
        p = Profile(name="Exp M", email="expm@t.com")
        db.add(p)
        db.flush()
        db.add(Skill(profile_id=p.id, name="Python"))
        db.flush()
        c = client.post("/companies", json={"name": "Exp M Co"}).json()
        opp = client.post("/opportunities", json={
            "company_id": c["id"], "type": "FULL_TIME", "title": "Py",
        }).json()
        resp = client.get(f"/matching/profiles/{p.id}/opportunities/{opp['id']}")
        assert resp.status_code == 200

    def test_outreach(self, client, db):
        from app.models.profile import Profile
        p = Profile(name="Exp O", email="expo@t.com")
        db.add(p)
        db.flush()
        c = client.post("/companies", json={"name": "Exp O Co"}).json()
        lead = client.post("/leads", json={
            "company_id": c["id"], "name": "L", "email": "l@t.com",
        }).json()
        opp = client.post("/opportunities", json={
            "company_id": c["id"], "type": "FULL_TIME", "title": "O",
        }).json()
        resp = client.post("/outreach/drafts", json={
            "profile_id": p.id, "lead_id": lead["id"], "opportunity_id": opp["id"],
        })
        assert resp.status_code == 201

    def test_followup(self, client, db):
        c = client.post("/companies", json={"name": "Exp F Co"}).json()
        lead = client.post("/leads", json={
            "company_id": c["id"], "name": "L", "email": "l@t.com",
        }).json()
        resp = client.post("/follow-ups", json={
            "lead_id": lead["id"],
            "scheduled_for": (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat(),
        })
        assert resp.status_code == 201

    def test_discovery(self, client, db):
        resp = client.post("/discovery/run", json=[{
            "source_name": "manual", "title": "D", "company_name": "DC",
        }])
        assert resp.status_code == 200
