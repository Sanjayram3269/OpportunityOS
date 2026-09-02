"""Comprehensive tests for milestone: Application Timeline UI + Campaign Drill-Down + Excel Export.

Tests cover:
  - Application Timeline export in Excel
  - Campaign drill-down analytics API
  - Campaign drill-down correctness (overview, conversion, activity, planning)
  - Timeline endpoint edge cases
  - Empty database handling
  - Division by zero safety
  - Regression for existing functionality
"""

from __future__ import annotations

from datetime import datetime, timedelta, timezone

import pytest
from openpyxl import load_workbook
from io import BytesIO

from app.models.application import Application
from app.models.application_event import ApplicationEvent
from app.models.campaign import Campaign
from app.models.campaign_opportunity import CampaignOpportunity
from app.models.company import Company
from app.models.followup import FollowUp
from app.models.message import Message
from app.models.opportunity import Opportunity
from app.export.workbook import build_workbook
from app.services.export import build_export_data
from app.services.application import create_application, transition_application


# ── Helpers ──────────────────────────────────────────────────────────────


def _create_company(db, name="TestCo"):
    company = Company(name=name, location="Bengaluru")
    db.add(company)
    db.flush()
    return company


def _create_opportunity(db, company, **kwargs):
    defaults = {
        "type": "INTERNSHIP",
        "title": "Test Intern",
        "status": "DISCOVERED",
        "priority": "HIGH",
        "match_score": 85,
    }
    defaults.update(kwargs)
    opp = Opportunity(company_id=company.id, **defaults)
    db.add(opp)
    db.flush()
    return opp


def _create_lead(db, company, name="TestLead"):
    lead = Lead(company_id=company.id, name=name, email=f"{name.lower()}@test.com")
    from app.models.lead import Lead
    lead = Lead(company_id=company.id, name=name, email=f"{name.lower()}@test.com")
    db.add(lead)
    db.flush()
    return lead


def _create_application_with_events(db, opp, statuses=None):
    """Create an application and transition through a series of statuses."""
    app = create_application(db, opportunity_id=opp.id)
    if statuses:
        for status in statuses:
            transition_application(db, app.id, status)
    db.flush()
    return app


# ══════════════════════════════════════════════════════════════════════════
# 1. APPLICATION TIMELINE EXPORT
# ══════════════════════════════════════════════════════════════════════════


class TestApplicationTimelineExport:
    def test_timeline_sheet_exists(self, db):
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        assert "App Timeline" in wb.sheetnames
        wb.close()

    def test_timeline_sheet_empty_when_no_events(self, db):
        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        assert rows == []
        assert "Application ID" in headers
        assert "Event Type" in headers

    def test_timeline_sheet_has_headers(self, db):
        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        expected_headers = [
            "Application ID", "Opportunity", "Company", "Event Type",
            "From Status", "To Status", "Label", "Occurred At",
        ]
        for h in expected_headers:
            assert h in headers

    def test_timeline_includes_creation_event(self, db):
        company = _create_company(db, "TimelineCo")
        opp = _create_opportunity(db, company, title="Timeline Opp")
        _create_application_with_events(db, opp)

        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        assert len(rows) >= 1
        event_types = [r[headers.index("Event Type")] for r in rows]
        assert "APPLICATION_CREATED" in event_types

    def test_timeline_includes_status_changes(self, db):
        company = _create_company(db, "StatusCo")
        opp = _create_opportunity(db, company, title="Status Opp")
        _create_application_with_events(db, opp, statuses=["READY", "APPLIED", "INTERVIEW"])

        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        event_types = [r[headers.index("Event Type")] for r in rows]
        assert "APPLICATION_CREATED" in event_types
        assert "STATUS_CHANGED" in event_types
        assert "APPLICATION_SUBMITTED" in event_types
        assert "INTERVIEW" in event_types

    def test_timeline_shows_opportunity_title(self, db):
        company = _create_company(db, "TitleCo")
        opp = _create_opportunity(db, company, title="Specific Title Here")
        _create_application_with_events(db, opp, statuses=["READY"])

        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        titles = [r[headers.index("Opportunity")] for r in rows]
        assert "Specific Title Here" in titles

    def test_timeline_shows_company_name(self, db):
        company = _create_company(db, "CompanyTimeline")
        opp = _create_opportunity(db, company)
        _create_application_with_events(db, opp, statuses=["READY"])

        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        companies = [r[headers.index("Company")] for r in rows]
        assert "CompanyTimeline" in companies

    def test_timeline_chronological_order(self, db):
        company = _create_company(db, "ChronoCo")
        opp = _create_opportunity(db, company)
        _create_application_with_events(db, opp, statuses=["READY", "APPLIED", "INTERVIEW", "OFFER"])

        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        timestamps = [r[headers.index("Occurred At")] for r in rows]
        # All timestamps should be in chronological order
        for i in range(len(timestamps) - 1):
            assert timestamps[i] <= timestamps[i + 1]

    def test_timeline_from_status_populated(self, db):
        company = _create_company(db, "FromCo")
        opp = _create_opportunity(db, company)
        _create_application_with_events(db, opp, statuses=["READY", "APPLIED"])

        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        # First event (APPLICATION_CREATED) has no from_status
        first = rows[0]
        assert first[headers.index("From Status")] == ""
        # Second event (STATUS_CHANGED for READY) has from_status
        second = rows[1]
        assert second[headers.index("From Status")] == "NOT_APPLIED"

    def test_timeline_to_status_populated(self, db):
        company = _create_company(db, "ToCo")
        opp = _create_opportunity(db, company)
        _create_application_with_events(db, opp, statuses=["READY"])

        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        # APPLICATION_CREATED: to_status = NOT_APPLIED
        assert rows[0][headers.index("To Status")] == "NOT_APPLIED"
        # STATUS_CHANGED: to_status = READY
        assert rows[1][headers.index("To Status")] == "READY"

    def test_timeline_multiple_applications(self, db):
        company = _create_company(db, "MultiCo")
        opp1 = _create_opportunity(db, company, title="Opp 1")
        opp2 = _create_opportunity(db, company, title="Opp 2")
        _create_application_with_events(db, opp1, statuses=["READY"])
        _create_application_with_events(db, opp2, statuses=["READY", "APPLIED"])

        data = build_export_data(db)
        headers, rows = data["application_timeline"]
        assert len(rows) >= 4  # 2 events for opp1 + 3 for opp2

    def test_timeline_in_workbook(self, db):
        company = _create_company(db, "WBCo")
        opp = _create_opportunity(db, company)
        _create_application_with_events(db, opp, statuses=["READY", "APPLIED"])

        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["App Timeline"]
        headers = [cell.value for cell in ws[1]]
        assert "Application ID" in headers
        assert ws.max_row >= 3  # header + 2 events
        wb.close()

    def test_timeline_empty_workbook(self, db):
        data = build_export_data(db)
        buf = build_workbook(data)
        wb = load_workbook(buf)
        ws = wb["App Timeline"]
        assert ws.max_row == 1  # header only
        wb.close()

    def test_timeline_api_endpoint(self, client, db):
        company = _create_company(db, "APICo")
        opp = _create_opportunity(db, company)
        app = _create_application_with_events(db, opp, statuses=["READY", "APPLIED"])

        resp = client.get(f"/applications/{app.id}/timeline")
        assert resp.status_code == 200
        body = resp.json()
        assert body["application_id"] == app.id
        assert body["current_status"] == "APPLIED"
        assert len(body["events"]) >= 3  # created + READY + APPLIED

    def test_timeline_api_not_found(self, client, db):
        resp = client.get("/applications/99999/timeline")
        assert resp.status_code == 404

    def test_timeline_api_chronological(self, client, db):
        company = _create_company(db, "ChronoAPICo")
        opp = _create_opportunity(db, company)
        app = _create_application_with_events(db, opp, statuses=["READY", "APPLIED", "INTERVIEW"])

        resp = client.get(f"/applications/{app.id}/timeline")
        body = resp.json()
        events = body["events"]
        assert len(events) >= 4
        for i in range(len(events) - 1):
            assert events[i]["occurred_at"] <= events[i + 1]["occurred_at"]


# ══════════════════════════════════════════════════════════════════════════
# 2. CAMPAIGN DRILL-DOWN ANALYTICS
# ══════════════════════════════════════════════════════════════════════════


class TestCampaignDrilldown:
    def test_drilldown_not_found(self, client, db):
        resp = client.get("/analytics/campaigns/99999")
        assert resp.status_code == 404

    def test_drilldown_empty_campaign(self, client, db):
        campaign = Campaign(name="Empty Campaign", type="FULL_TIME", status="ACTIVE")
        db.add(campaign)
        db.flush()

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        assert resp.status_code == 200
        body = resp.json()
        assert body["campaign_id"] == campaign.id
        assert body["campaign_name"] == "Empty Campaign"
        assert body["campaign_status"] == "ACTIVE"
        assert body["overview"]["total_opportunities"] == 0

    def test_drilldown_with_opportunities(self, client, db):
        company = _create_company(db, "DrillCo")
        campaign = Campaign(name="Drill Campaign", type="INTERNSHIP", status="ACTIVE")
        db.add(campaign)
        db.flush()

        opp = _create_opportunity(db, company, match_score=90)
        co = CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp.id)
        db.add(co)
        db.flush()

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        assert resp.status_code == 200
        body = resp.json()
        assert body["overview"]["total_opportunities"] == 1
        assert body["overview"]["high_match"] == 1

    def test_drilldown_conversion_rates(self, client, db):
        company = _create_company(db, "ConvCo")
        campaign = Campaign(name="Conv Campaign", type="INTERNSHIP", status="ACTIVE")
        db.add(campaign)
        db.flush()

        opp = _create_opportunity(db, company)
        co = CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp.id)
        db.add(co)
        db.flush()

        _create_application_with_events(db, opp, statuses=["READY", "APPLIED", "INTERVIEW"])

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        body = resp.json()
        assert body["overview"]["applications_submitted"] == 1
        assert body["overview"]["interviews"] == 1
        assert body["conversion"]["application_rate"] is not None
        assert body["conversion"]["interview_rate"] is not None

    def test_drilldown_conversion_zero_denominator(self, client, db):
        company = _create_company(db, "ZeroCo")
        campaign = Campaign(name="Zero Campaign", type="INTERNSHIP", status="ACTIVE")
        db.add(campaign)
        db.flush()

        opp = _create_opportunity(db, company)
        co = CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp.id)
        db.add(co)
        db.flush()
        # Create application in READY state (not submitted)
        _create_application_with_events(db, opp, statuses=["READY"])

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        body = resp.json()
        # apps_started=1 (READY), apps_submitted=0 → application_rate=1.0 (1/1)
        # But no submitted apps → interview_rate should be None
        assert body["conversion"]["interview_rate"] is None
        assert body["overview"]["applications_started"] == 1
        assert body["overview"]["applications_submitted"] == 0

    def test_drilldown_planning_distribution(self, client, db):
        company = _create_company(db, "PlanCo")
        campaign = Campaign(name="Plan Campaign", type="INTERNSHIP", status="ACTIVE")
        db.add(campaign)
        db.flush()

        # Summer 2027 opportunity
        opp_s27 = _create_opportunity(
            db, company, title="S27",
            deadline=datetime(2027, 5, 15, tzinfo=timezone.utc),
        )
        db.add(CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp_s27.id))

        # Unknown (no deadline) opportunity
        opp_unknown = _create_opportunity(db, company, title="Unknown DL")
        db.add(CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp_unknown.id))
        db.flush()

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        body = resp.json()
        assert body["planning"]["SUMMER_2027"] == 1
        assert body["planning"]["UNKNOWN"] == 1

    def test_drilldown_activity_counts(self, client, db):
        company = _create_company(db, "ActCo")
        campaign = Campaign(name="Act Campaign", type="INTERNSHIP", status="ACTIVE")
        db.add(campaign)
        db.flush()

        opp = _create_opportunity(db, company)
        co = CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp.id)
        db.add(co)
        db.flush()

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        body = resp.json()
        assert body["activity"]["open_actions"] == 0
        assert body["activity"]["overdue_actions"] == 0
        assert body["activity"]["outreach_pending_approval"] == 0
        assert body["activity"]["followups_due"] == 0

    def test_drilldown_outreach_activity(self, client, db):
        from app.models.lead import Lead
        company = _create_company(db, "MsgCo")
        campaign = Campaign(name="Msg Campaign", type="INTERNSHIP", status="ACTIVE")
        db.add(campaign)
        db.flush()

        opp = _create_opportunity(db, company)
        db.add(CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp.id))

        lead = Lead(company_id=company.id, name="MsgLead", email="msg@test.com")
        db.add(lead)
        db.flush()

        msg = Message(
            lead_id=lead.id, opportunity_id=opp.id,
            channel="EMAIL", direction="OUTBOUND",
            body="Test", status="PENDING_APPROVAL",
        )
        db.add(msg)
        db.flush()

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        body = resp.json()
        assert body["activity"]["outreach_pending_approval"] == 1

    def test_drilldown_overview_status_breakdown(self, client, db):
        company = _create_company(db, "BreakCo")
        campaign = Campaign(name="Break Campaign", type="INTERNSHIP", status="ACTIVE")
        db.add(campaign)
        db.flush()

        opp1 = _create_opportunity(db, company, title="Opp 1")
        opp2 = _create_opportunity(db, company, title="Opp 2")
        db.add(CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp1.id))
        db.add(CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp2.id))
        db.flush()

        _create_application_with_events(db, opp1, statuses=["READY", "APPLIED"])
        _create_application_with_events(db, opp2, statuses=["READY"])

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        body = resp.json()
        assert body["overview"]["applications_started"] >= 1
        assert body["overview"]["applications_submitted"] >= 1

    def test_drilldown_campaign_info(self, client, db):
        campaign = Campaign(name="Info Campaign", type="RESEARCH", status="PAUSED")
        db.add(campaign)
        db.flush()

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        body = resp.json()
        assert body["campaign_name"] == "Info Campaign"
        assert body["campaign_status"] == "PAUSED"

    def test_drilldown_full_lifecycle(self, client, db):
        company = _create_company(db, "FullCo")
        campaign = Campaign(name="Full Campaign", type="INTERNSHIP", status="ACTIVE")
        db.add(campaign)
        db.flush()

        opp = _create_opportunity(db, company, match_score=95)
        db.add(CampaignOpportunity(campaign_id=campaign.id, opportunity_id=opp.id))
        db.flush()

        _create_application_with_events(
            db, opp,
            statuses=["READY", "APPLIED", "ASSESSMENT", "INTERVIEW", "FINAL_ROUND", "OFFER"],
        )

        resp = client.get(f"/analytics/campaigns/{campaign.id}")
        body = resp.json()
        assert body["overview"]["total_opportunities"] == 1
        assert body["overview"]["high_match"] == 1
        assert body["overview"]["applications_submitted"] == 1
        # Final status is OFFER, so counts by current status:
        assert body["overview"]["interviews"] == 0  # current status is OFFER, not INTERVIEW
        assert body["overview"]["final_rounds"] == 0  # current status is OFFER, not FINAL_ROUND
        assert body["overview"]["offers"] == 1  # current status IS OFFER
        assert body["conversion"]["application_rate"] == 1.0
        # interview_rate = (INTERVIEW + FINAL_ROUND) / submitted = 0/1 = 0.0
        # (current status is OFFER, not INTERVIEW/FINAL_ROUND)
        assert body["conversion"]["interview_rate"] == 0.0
        assert body["conversion"]["offer_rate"] == 1.0


# ══════════════════════════════════════════════════════════════════════════
# 3. TIMELINE EDGE CASES
# ══════════════════════════════════════════════════════════════════════════


class TestTimelineEdgeCases:
    def test_empty_timeline(self, client, db):
        company = _create_company(db, "EdgeCo")
        opp = _create_opportunity(db, company)
        app = create_application(db, opportunity_id=opp.id)
        db.flush()

        resp = client.get(f"/applications/{app.id}/timeline")
        assert resp.status_code == 200
        body = resp.json()
        # Should have at least the APPLICATION_CREATED event
        assert len(body["events"]) >= 1
        assert body["events"][0]["event_type"] == "APPLICATION_CREATED"

    def test_timeline_with_multiple_transitions(self, client, db):
        company = _create_company(db, "MultiTransCo")
        opp = _create_opportunity(db, company)
        app = _create_application_with_events(
            db, opp,
            statuses=["READY", "APPLIED", "INTERVIEW", "REJECTED"],
        )
        db.flush()

        resp = client.get(f"/applications/{app.id}/timeline")
        body = resp.json()
        assert body["current_status"] == "REJECTED"
        event_types = [e["event_type"] for e in body["events"]]
        assert "APPLICATION_CREATED" in event_types
        assert "APPLICATION_SUBMITTED" in event_types
        assert "INTERVIEW" in event_types
        assert "REJECTED" in event_types

    def test_timeline_withdrawn(self, client, db):
        company = _create_company(db, "WithCo")
        opp = _create_opportunity(db, company)
        app = _create_application_with_events(
            db, opp,
            statuses=["READY", "APPLIED", "WITHDRAWN"],
        )
        db.flush()

        resp = client.get(f"/applications/{app.id}/timeline")
        body = resp.json()
        assert body["current_status"] == "WITHDRAWN"
        event_types = [e["event_type"] for e in body["events"]]
        assert "WITHDRAWN" in event_types

    def test_timeline_labels_meaningful(self, client, db):
        company = _create_company(db, "LabelCo")
        opp = _create_opportunity(db, company)
        app = _create_application_with_events(
            db, opp,
            statuses=["READY", "APPLIED", "INTERVIEW"],
        )
        db.flush()

        resp = client.get(f"/applications/{app.id}/timeline")
        body = resp.json()
        for event in body["events"]:
            assert event["label"]  # Non-empty
            assert len(event["label"]) > 3  # Meaningful label

    def test_no_fabricated_backfill(self, client, db):
        """Old applications should not have fabricated history."""
        company = _create_company(db, "BackfillCo")
        opp = _create_opportunity(db, company)
        # Create application directly (not through create_application to simulate old record)
        app = Application(
            opportunity_id=opp.id,
            status="APPLIED",
            applied_at=datetime.now(timezone.utc),
        )
        db.add(app)
        db.flush()

        resp = client.get(f"/applications/{app.id}/timeline")
        body = resp.json()
        # Old record — may have 0 events (no fabricated backfill)
        # Or only APPLICATION_CREATED if create_application was used
        for event in body["events"]:
            assert event["event_type"] in [
                "APPLICATION_CREATED", "STATUS_CHANGED",
                "APPLICATION_SUBMITTED", "ASSESSMENT", "INTERVIEW",
                "FINAL_ROUND", "OFFER", "ACCEPTED", "REJECTED", "WITHDRAWN",
            ]


# ══════════════════════════════════════════════════════════════════════════
# 4. REGRESSION
# ══════════════════════════════════════════════════════════════════════════


class TestExistingRegression:
    def test_export_still_works(self, client, db):
        resp = client.get("/exports/opportunities.xlsx")
        assert resp.status_code == 200

    def test_analytics_still_works(self, client, db):
        resp = client.get("/analytics/overview")
        assert resp.status_code == 200

    def test_dashboard_still_works(self, client, db):
        resp = client.get("/dashboard/overview")
        assert resp.status_code == 200

    def test_applications_still_work(self, client, db):
        resp = client.get("/applications")
        assert resp.status_code == 200

    def test_actions_still_work(self, client, db):
        resp = client.get("/actions")
        assert resp.status_code == 200

    def test_campaigns_still_work(self, client, db):
        resp = client.get("/campaigns")
        assert resp.status_code == 200

    def test_planning_still_works(self, client, db):
        resp = client.get("/opportunities/planning")
        assert resp.status_code == 200
